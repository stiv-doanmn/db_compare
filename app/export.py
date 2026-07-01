"""Phase 5 export — Excel (.xlsx, gói .zip) và CSV báo cáo so sánh.

Toàn bộ bản ghi sai lệch nằm ở spill file trên đĩa (xem app/spill.py); export
ĐỌC STREAM từ đó nên RAM chỉ giữ 1 workbook ≤ MAX_SAMPLES dòng tại một thời điểm.

Excel: report CHIA THÀNH NHIỀU FILE .xlsx (mỗi file ≤ MAX_SAMPLES dòng) gói .zip:
- File đầu: sheet "Tổng quan" + sheet "Constraint Summary" + các sheet bảng.
- Mỗi bảng 1 sheet riêng; bảng có quá nhiều dòng diff bị cắt thành nhiều sheet
  "(2)", "(3)"… và tràn sang file kế khi vượt MAX_SAMPLES (vd 800k → 500k+300k).
CSV: 1 file phẳng, mỗi dòng là 1 phát hiện (finding) — dễ import/pivot.
"""
from __future__ import annotations

import csv
import io
import zipfile

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.worksheet.worksheet import Worksheet

from . import spill
from .config import MAX_SAMPLES

# ----- Bảng màu chuẩn doanh nghiệp ----------------------------------------- #
_NAVY = "181D26"
_WHITE = "FFFFFF"
_GREY_SOFT = "F1F3F5"
_GREY_LINE = "DDDDDD"

_FILL_HEAD = PatternFill("solid", fgColor=_NAVY)
_FILL_SUBHEAD = PatternFill("solid", fgColor="E0E2E6")
_FILL_BAND = PatternFill("solid", fgColor=_GREY_SOFT)
_FILL_OK = PatternFill("solid", fgColor="E7F0E7")
_FILL_WARN = PatternFill("solid", fgColor="F7ECD2")
_FILL_ERR = PatternFill("solid", fgColor="F7E6DF")
_FILL_A = PatternFill("solid", fgColor="FBEAE3")  # giá trị bên A
_FILL_B = PatternFill("solid", fgColor="E7EEFB")  # giá trị bên B

_FONT_HEAD = Font(name="Calibri", size=11, bold=True, color=_WHITE)
_FONT_TITLE = Font(name="Calibri", size=14, bold=True, color=_NAVY)
_FONT_BOLD = Font(name="Calibri", size=11, bold=True, color=_NAVY)
_FONT_OK = Font(name="Calibri", size=11, bold=True, color="006400")
_FONT_WARN = Font(name="Calibri", size=11, bold=True, color="8A5A00")
_FONT_ERR = Font(name="Calibri", size=11, bold=True, color="AA2D00")
_FONT_MUTED = Font(name="Calibri", size=10, color="6B7280")
_FONT_LINK = Font(name="Calibri", size=11, bold=True, color="1B61C9", underline="single")
_FONT_BACK = Font(name="Calibri", size=11, bold=True, color=_NAVY, underline="single")
_FILL_BACK = PatternFill("solid", fgColor="F4D35E")  # nút back nổi bật

_SUMMARY_TITLE = "Tổng quan"
_CONSTRAINT_TITLE = "Constraint Summary"
_KEYWORD_TITLE = "Tìm kiếm từ khóa"
_UNSET = object()  # sentinel cho group-id ban đầu (sọc màu record diff)

_thin = Side(style="thin", color=_GREY_LINE)
_BORDER = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
_RIGHT = Alignment(horizontal="right", vertical="center")

_STATUS_FILL = {"done": _FILL_OK, "warning": _FILL_WARN, "error": _FILL_ERR}
_STATUS_FONT = {"done": _FONT_OK, "warning": _FONT_WARN, "error": _FONT_ERR}


def _status_label(p) -> str:
    return {
        "done": "OK — Khớp",
        "warning": "Khác biệt",
        "error": "Lỗi",
    }.get(p.status, p.status)


# --------------------------------------------------------------------------- #
# Excel — helpers
# --------------------------------------------------------------------------- #
def _header_row(ws: Worksheet, row: int, headers: list[str]) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.fill = _FILL_HEAD
        cell.font = _FONT_HEAD
        cell.alignment = _CENTER
        cell.border = _BORDER


def _safe_sheet_title(name: str, used: set[str]) -> str:
    # Excel cấm : \ / ? * [ ] và tối đa 31 ký tự, không trùng.
    clean = "".join("_" if ch in r":\/?*[]" else ch for ch in name)[:31] or "table"
    title, i = clean, 1
    while title.lower() in used:
        suffix = f"~{i}"
        title = clean[: 31 - len(suffix)] + suffix
        i += 1
    used.add(title.lower())
    return title


# --------------------------------------------------------------------------- #
# Excel — record diff rows (stream từ spill; mỗi row: list[(value, fill|None)])
# --------------------------------------------------------------------------- #
def _table_rowcount(p) -> int:
    """Tổng số dòng record-diff của 1 bảng (tính từ counters, không đọc spill)."""
    if p.mode == "count-only":
        return 0
    return (p.only_in_a or 0) + (p.only_in_b or 0) + (p.mismatch_row_count or 0)


def _iter_record_rows(job, p):
    """Yield (group_id, cells) cho từng dòng record-diff, stream từ spill.

    group_id để export tô màu sọc trắng/xám theo NHÓM: các dòng cùng 1 id (vd 1
    bản ghi mismatch có nhiều cột khác) cùng màu; id kế bên đổi màu.
    """
    pair = job.pair_key()
    la, lb = job.label_a, job.label_b

    for key in spill.iter_records(pair, p.name, "only_a"):
        yield ("a", key), [f"Chỉ có ở {la}", str(key), None, None, None]
    for key in spill.iter_records(pair, p.name, "only_b"):
        yield ("b", key), [f"Chỉ có ở {lb}", str(key), None, None, None]
    for rec in spill.iter_records(pair, p.name, "mismatch_detail"):
        rid = rec.get("id")
        diffs = rec.get("diffs") or []
        if diffs:
            for d in diffs:
                yield ("m", rid), ["Cùng id, khác giá trị", str(rid),
                                   d.get("col"), d.get("a"), d.get("b")]
        else:  # không so được cột (1 bên thiếu id) → vẫn liệt kê id
            yield ("m", rid), ["Cùng id, khác giá trị", str(rid), None, None, None]


# --------------------------------------------------------------------------- #
# Excel — sheet Tổng quan
# --------------------------------------------------------------------------- #
def _build_summary(ws: Worksheet, job, titles: dict, first_part: dict) -> None:
    """titles: table -> sheet title (chỉ bảng có sheet trong CÙNG file → link nội bộ).
    first_part: table -> chỉ số file (0-based) chứa sheet đầu của bảng đó."""
    ws.sheet_view.showGridLines = False
    c = job.counters()

    ws["A1"] = "BÁO CÁO SO SÁNH DATABASE"
    ws["A1"].font = _FONT_TITLE
    ws["A2"] = f"{job.label_a} ({job.version_a})  ↔  {job.label_b} ({job.version_b})"
    ws["A2"].font = _FONT_MUTED
    ws["A3"] = (
        f"Tổng: {c['total']}   •   Khớp: {c['done']}   •   "
        f"Khác biệt: {c['warning']}   •   Lỗi: {c['error']}"
    )
    ws["A3"].font = _FONT_BOLD

    head = [
        "Bảng", "Mode", "Trạng thái",
        f"Rows {job.label_a}", f"Rows {job.label_b}", "Δ rows",
        f"Chỉ ở {job.label_a}", f"Chỉ ở {job.label_b}", "Giá trị khác",
        "Ghi chú",
    ]
    hr = 5
    _header_row(ws, hr, head)

    r = hr + 1
    for p in job.progress.values():
        name_cell = ws.cell(row=r, column=1, value=p.name)
        title = titles.get(p.name)
        if title:  # sheet nằm cùng file → hyperlink nội bộ
            name_cell.hyperlink = Hyperlink(
                ref=name_cell.coordinate, location=f"'{title}'!A1", display=p.name
            )
            name_cell.font = _FONT_LINK
        elif first_part.get(p.name):  # sheet ở file phần khác (>0)
            name_cell.value = f"{p.name}  (xem file phần {first_part[p.name] + 1})"
            name_cell.font = _FONT_BOLD
        else:
            name_cell.font = _FONT_BOLD
        ws.cell(row=r, column=2, value=p.mode)
        st = ws.cell(row=r, column=3, value=_status_label(p))
        st.fill = _STATUS_FILL.get(p.status, _FILL_BAND)
        st.font = _STATUS_FONT.get(p.status, _FONT_BOLD)
        st.alignment = _CENTER
        ws.cell(row=r, column=4, value=p.count_a).alignment = _RIGHT
        ws.cell(row=r, column=5, value=p.count_b).alignment = _RIGHT
        dcell = ws.cell(row=r, column=6, value=p.count_delta)
        dcell.alignment = _RIGHT
        if p.count_delta != 0:
            dcell.font = _FONT_ERR
        ws.cell(row=r, column=7, value=p.only_in_a).alignment = _RIGHT
        ws.cell(row=r, column=8, value=p.only_in_b).alignment = _RIGHT
        ws.cell(row=r, column=9, value=p.value_mismatch).alignment = _RIGHT
        note = p.note
        if p.error:
            note = (note + " | " if note else "") + p.error
        ws.cell(row=r, column=10, value=note).alignment = _LEFT
        for col in range(1, 11):
            cell = ws.cell(row=r, column=col)
            cell.border = _BORDER
            if r % 2 == 0 and col != 3:
                if cell.fill.fgColor.rgb in (None, "00000000"):
                    cell.fill = _FILL_BAND
        r += 1

    widths = [40, 13, 14, 16, 16, 12, 16, 16, 14, 50]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    ws.auto_filter.ref = f"A{hr}:J{r - 1}"


# --------------------------------------------------------------------------- #
# Excel — sheet tổng hợp thay đổi constraint
# --------------------------------------------------------------------------- #
_CONSTRAINT_HEAD = [
    "No.", "Object Type", "Name", "Change Type", "Description", "Related Issues",
    "Solution Temp", "Final Solution", "Status", "Solution Description", "Module",
    "Used in AstraLink", "Astralink module", "File / Line",
]
# kind (ConstraintDiff) -> Change Type
_CHANGE_TYPE = {"added": "ADDED", "dropped": "REMOVED", "changed": "SCHEMA_CHANGED"}


def _build_constraint_summary(ws: Worksheet, job) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "TỔNG HỢP THAY ĐỔI CONSTRAINT"
    ws["A1"].font = _FONT_TITLE
    ws["A2"] = f"{job.label_a}  →  {job.label_b}"
    ws["A2"].font = _FONT_MUTED

    hr = 4
    _header_row(ws, hr, _CONSTRAINT_HEAD)

    r = hr + 1
    no = 1
    schema_tables = getattr(job, "schema_tables", {}) or {}
    for tname, td in schema_tables.items():
        for c in getattr(td, "constraints", []):
            desc = f"[{tname}] {c.type_label}"
            if c.def_a or c.def_b:
                desc += (
                    f" · {job.label_a}: {c.def_a or '∅'}"
                    f" | {job.label_b}: {c.def_b or '∅'}"
                )
            values = [
                no,                                          # No.
                "CONSTRAINTS",                               # Object Type
                c.name,                                      # Name
                _CHANGE_TYPE.get(c.kind, c.kind.upper()),    # Change Type
                desc,                                        # Description
                "", "", "", "", "", "", "", "", "",          # các cột để trống
            ]
            for col, val in enumerate(values, start=1):
                cell = ws.cell(row=r, column=col, value=val)
                cell.border = _BORDER
                cell.alignment = _LEFT if col not in (1,) else _CENTER
            r += 1
            no += 1

    if no == 1:  # không có constraint nào lệch
        ws.cell(row=r, column=1, value="Không có thay đổi constraint.").font = _FONT_OK

    widths = [6, 14, 30, 16, 60, 14, 14, 14, 12, 22, 16, 16, 18, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    if no > 1:
        ws.auto_filter.ref = f"A{hr}:N{r - 1}"


# --------------------------------------------------------------------------- #
# Excel — sheet kết quả tìm cụm từ khóa (dò ở DB B)
# --------------------------------------------------------------------------- #
_KEYWORD_HEAD = [
    "No.", "Loại", "Từ khóa / Mẫu", "Bảng", "Database", "Số dòng khớp",
    "Cột đã tìm", "id mẫu", "Câu query",
]
_KIND_LABEL = {"keyword": "từ khóa", "pattern": "mẫu"}


def _fmt_cols_list(cols: list[str], limit: int = 12) -> str:
    if not cols:
        return ""
    if len(cols) <= limit:
        return ", ".join(cols)
    return ", ".join(cols[:limit]) + f" … (+{len(cols) - limit} cột)"


def _build_keyword_search(ws: Worksheet, job) -> None:
    ws.sheet_view.showGridLines = False
    ws["A1"] = "KẾT QUẢ TÌM CỤM TỪ KHÓA"
    ws["A1"].font = _FONT_TITLE
    ws["A2"] = (
        f"Dò trong cột kiểu text của {job.label_b} · không phân biệt hoa/thường "
        f"(keyword: ILIKE · mẫu: regex ~*)"
    )
    ws["A2"].font = _FONT_MUTED
    from .config import SEARCH_PATTERN_MAP
    pat_labels = [SEARCH_PATTERN_MAP[k]["label"] for k in getattr(job, "search_patterns", []) if k in SEARCH_PATTERN_MAP]
    parts = []
    if job.keywords:
        parts.append("Từ khóa: " + "  ·  ".join(job.keywords))
    if pat_labels:
        parts.append("Mẫu: " + "  ·  ".join(pat_labels))
    ws["A3"] = "   |   ".join(parts) if parts else "(chưa nhập từ khóa / bật mẫu)"
    ws["A3"].font = _FONT_BOLD

    hr = 5
    _header_row(ws, hr, _KEYWORD_HEAD)

    r = hr + 1
    no = 1
    for h in job.keyword_hits:
        kind = _KIND_LABEL.get(h.kind, h.kind)
        if h.error:
            ws.cell(row=r, column=1, value=no)
            ws.cell(row=r, column=2, value=kind)
            ws.cell(row=r, column=3, value=h.keyword)
            ws.cell(row=r, column=4, value=h.table)
            ws.cell(row=r, column=5, value=h.db_label)
            ecell = ws.cell(row=r, column=6, value="lỗi")
            ecell.font = _FONT_ERR
            ws.cell(row=r, column=9, value=(h.query or "") + f"  — {h.error}")
        else:
            ws.cell(row=r, column=1, value=no)
            ws.cell(row=r, column=2, value=kind)
            ws.cell(row=r, column=3, value=h.keyword).font = _FONT_BOLD
            ws.cell(row=r, column=4, value=h.table)
            ws.cell(row=r, column=5, value=h.db_label)
            mc = ws.cell(row=r, column=6, value=h.match_count)
            mc.alignment = _RIGHT
            mc.font = _FONT_WARN
            ws.cell(row=r, column=7, value=_fmt_cols_list(h.columns)).alignment = _LEFT
            ids = ", ".join(str(i) for i in h.sample_ids)
            if h.has_id and h.match_count > len(h.sample_ids):
                ids += " …"
            ws.cell(row=r, column=8, value=ids or ("(bảng không có id)" if not h.has_id else "")).alignment = _LEFT
            ws.cell(row=r, column=9, value=h.query).alignment = _LEFT
        for col in range(1, 10):
            ws.cell(row=r, column=col).border = _BORDER
        r += 1
        no += 1

    if no == 1:
        has_terms = job.keywords or getattr(job, "search_patterns", None)
        msg = ("Không tìm thấy dữ liệu khớp từ khóa/mẫu nào."
               if has_terms else "Chưa nhập từ khóa/bật mẫu ở bước Config.")
        ws.cell(row=r, column=1, value=msg).font = _FONT_OK

    widths = [6, 12, 26, 30, 12, 13, 42, 36, 90]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    if no > 1:
        ws.auto_filter.ref = f"A{hr}:I{r - 1}"


# --------------------------------------------------------------------------- #
# Excel — sheet chi tiết 1 bảng (có thể là 1 phần của bảng bị cắt)
# --------------------------------------------------------------------------- #
def _build_table_sheet(
    ws: Worksheet, job, p, row_iter, max_rows, idx: int, total: int, summary_in_file: bool
) -> None:
    """Lấy tối đa max_rows dòng từ row_iter (None = lấy hết) ghi vào sheet này.
    idx/total: phần idx+1/total của bảng khi bị cắt nhiều sheet."""
    ws.sheet_view.showGridLines = False
    is_first = idx == 0

    if summary_in_file:
        back = ws.cell(row=1, column=6, value="← Về Tổng quan")
        back.hyperlink = Hyperlink(
            ref=back.coordinate, location=f"'{_SUMMARY_TITLE}'!A1", display="← Về Tổng quan"
        )
        back.font = _FONT_BACK
        back.fill = _FILL_BACK
        back.alignment = _CENTER
        back.border = _BORDER

    suffix = "" if total == 1 else f"  (phần {idx + 1}/{total})"
    ws["A1"] = p.name + suffix
    ws["A1"].font = _FONT_TITLE
    ws["A2"] = f"Mode: {p.mode}   •   Trạng thái: {_status_label(p)}"
    ws["A2"].font = _STATUS_FONT.get(p.status, _FONT_MUTED)

    if not is_first:
        # Sheet tiếp theo: chỉ ghi tiếp record-diff, không lặp lại metadata.
        ws["A3"] = "Tiếp theo danh sách bản ghi khác biệt từ phần trước."
        ws["A3"].font = _FONT_MUTED
        _write_record_block(ws, job, row_iter, max_rows, start_row=5, header_only_if_empty=False)
        for i, w in enumerate([28, 26, 30, 36, 36, 18], start=1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return

    if p.note:
        ws["A3"] = f"Ghi chú: {p.note}"
        ws["A3"].font = _FONT_MUTED
    if p.error:
        ws["A4"] = f"Lỗi: {p.error}"
        ws["A4"].font = _FONT_ERR

    # Khối row count
    r = 6
    ws.cell(row=r, column=1, value="Số lượng bản ghi").font = _FONT_BOLD
    r += 1
    _header_row(ws, r, ["", job.label_a, job.label_b, "Δ"])
    r += 1
    ws.cell(row=r, column=1, value="Rows").font = _FONT_BOLD
    ws.cell(row=r, column=2, value=p.count_a).alignment = _RIGHT
    ws.cell(row=r, column=3, value=p.count_b).alignment = _RIGHT
    dc = ws.cell(row=r, column=4, value=p.count_delta)
    dc.alignment = _RIGHT
    if p.count_delta:
        dc.font = _FONT_ERR
    for col in range(1, 5):
        ws.cell(row=r, column=col).border = _BORDER

    # Khối schema khác biệt
    sd = (getattr(job, "schema_tables", {}) or {}).get(p.name)
    if sd and getattr(sd, "columns", None):
        r += 3
        ws.cell(row=r, column=1, value="Schema khác biệt").font = _FONT_BOLD
        r += 1
        _header_row(ws, r, ["Cột", "Kiểu khác biệt", job.label_a, job.label_b, ""])
        r += 1
        start = r
        _KIND = {
            "added": f"thêm mới ở {job.label_b}",
            "dropped": f"bị bỏ ở {job.label_b}",
            "type_changed": "đổi kiểu",
            "nullable_changed": "đổi nullable",
            "default_changed": "đổi default",
        }
        for c in sd.columns:
            ws.cell(row=r, column=1, value=c.name)
            ws.cell(row=r, column=2, value=_KIND.get(c.kind, c.kind))
            va = ws.cell(row=r, column=3, value=(c.detail_a if c.detail_a is not None else "∅"))
            vb = ws.cell(row=r, column=4, value=(c.detail_b if c.detail_b is not None else "∅"))
            va.fill, vb.fill = _FILL_A, _FILL_B
            r += 1
        for rr in range(start, r):
            for col in range(1, 5):
                ws.cell(row=rr, column=col).border = _BORDER

    # Khối constraint khác biệt
    if sd and getattr(sd, "constraints", None):
        r += 3
        ws.cell(row=r, column=1, value="Constraint khác biệt").font = _FONT_BOLD
        r += 1
        _header_row(ws, r, ["Loại", "Tên constraint", "Khác biệt", job.label_a, job.label_b])
        r += 1
        start = r
        _CKIND = {
            "added": f"thêm mới ở {job.label_b}",
            "dropped": f"bị bỏ ở {job.label_b}",
            "changed": "đổi định nghĩa",
        }
        for c in sd.constraints:
            ws.cell(row=r, column=1, value=c.type_label)
            ws.cell(row=r, column=2, value=c.name)
            ws.cell(row=r, column=3, value=_CKIND.get(c.kind, c.kind))
            va = ws.cell(row=r, column=4, value=(c.def_a if c.def_a is not None else "∅"))
            vb = ws.cell(row=r, column=5, value=(c.def_b if c.def_b is not None else "∅"))
            va.fill, vb.fill = _FILL_A, _FILL_B
            r += 1
        for rr in range(start, r):
            for col in range(1, 6):
                ws.cell(row=rr, column=col).border = _BORDER

    # Khối record diff (chỉ khi không phải count-only)
    if p.mode != "count-only":
        r += 3
        _write_record_block(ws, job, row_iter, max_rows, start_row=r, header_only_if_empty=True)

    for i, w in enumerate([28, 26, 30, 36, 36, 18], start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_record_block(
    ws: Worksheet, job, row_iter, max_rows, start_row: int, header_only_if_empty: bool
) -> None:
    r = start_row
    ws.cell(row=r, column=1, value="Bản ghi khác biệt").font = _FONT_BOLD
    r += 1
    _header_row(ws, r, [
        "Loại khác biệt", "id / key", "Cột",
        f"Giá trị {job.label_a}", f"Giá trị {job.label_b}",
    ])
    r += 1
    start = r
    written = 0
    prev_gid = _UNSET
    stripe = False
    for gid, cells in row_iter:
        if gid != prev_gid:  # đổi nhóm id → lật màu sọc
            stripe = not stripe
            prev_gid = gid
        fill = _FILL_BAND if stripe else None  # xám / trắng
        for col, val in enumerate(cells, start=1):
            cell = ws.cell(row=r, column=col, value=val)
            if fill is not None:
                cell.fill = fill
        r += 1
        written += 1
        if max_rows is not None and written >= max_rows:
            break
    # Không kẻ border từng dòng record (có thể tới hàng trăm nghìn) — giữ tốc độ.
    if r == start and header_only_if_empty:
        ws.cell(row=r, column=1, value="Không có khác biệt theo bản ghi.").font = _FONT_OK


# --------------------------------------------------------------------------- #
# Excel — chia bảng thành các sheet/file (batch)
# --------------------------------------------------------------------------- #
def _plan_parts(job, limit: int) -> list[list[tuple]]:
    """Trả về list các file; mỗi file là list placement (p, idx, total, chunk_len).

    Kích thước tính từ counters (không đọc spill):
    - Mỗi bảng → ≥1 sheet; record-diff > limit bị cắt thành nhiều sheet.
    - Đóng gói các sheet vào file sao cho tổng dòng/file ≤ limit.
    """
    sheets: list[tuple] = []
    for p in job.progress.values():
        n = _table_rowcount(p)
        total = 1 if n == 0 else (n + limit - 1) // limit
        for i in range(total):
            chunk_len = 0 if n == 0 else min(limit, n - i * limit)
            sheets.append((p, i, total, chunk_len))

    parts: list[list[tuple]] = [[]]
    cur = 0
    for sh in sheets:
        chunk_len = sh[3]
        if cur > 0 and cur + chunk_len > limit:
            parts.append([])
            cur = 0
        parts[-1].append(sh)
        cur += chunk_len
    return parts


def build_workbook_zip(job) -> bytes:
    """Báo cáo Excel gói trong .zip — 1 hoặc nhiều file .xlsx tuỳ khối lượng.

    Mỗi file chứa tối đa MAX_SAMPLES dòng sai lệch; bảng vượt ngưỡng tràn sang
    file kế (vd 800k diff → file1 500k + file2 300k)."""
    parts = _plan_parts(job, MAX_SAMPLES)

    # File chứa sheet ĐẦU của mỗi bảng (để Tổng quan ghi link / chú thích).
    first_part: dict = {}
    for pi, part in enumerate(parts):
        for (p, idx, _total, _clen) in part:
            if idx == 0:
                first_part.setdefault(p.name, pi)

    job_tag = job.id[:8]
    files: list[tuple[str, bytes]] = []
    n_parts = len(parts)

    # Iterator record-diff dùng chung cho các sheet liên tiếp của CÙNG 1 bảng
    # (kể cả khi bảng tràn qua nhiều file) → đọc spill đúng 1 lần, tuần tự.
    cur_p = None
    row_iter = iter(())

    for pi, part in enumerate(parts):
        wb = Workbook()
        wb.remove(wb.active)  # bỏ sheet mặc định, tạo tường minh theo thứ tự
        used: set[str] = set()

        summary_ws = None
        if pi == 0:
            summary_ws = wb.create_sheet(_SUMMARY_TITLE)
            used.add(_SUMMARY_TITLE.lower())
            cs_ws = wb.create_sheet(_safe_sheet_title(_CONSTRAINT_TITLE, used))
            _build_constraint_summary(cs_ws, job)
            # Sheet tìm từ khóa — chỉ khi thực sự có chạy tìm (mode both/keyword).
            if getattr(job, "keywords", None) and getattr(job, "run_mode", "both") in ("both", "keyword"):
                ks_ws = wb.create_sheet(_safe_sheet_title(_KEYWORD_TITLE, used))
                _build_keyword_search(ks_ws, job)

        titles: dict = {}  # table -> sheet title (chỉ sheet đầu, trong file này)
        for (p, idx, total, clen) in part:
            base = p.name if total == 1 else f"{p.name} ({idx + 1})"
            title = _safe_sheet_title(base, used)
            ws = wb.create_sheet(title)
            if cur_p is not p:
                row_iter = _iter_record_rows(job, p) if p.mode != "count-only" else iter(())
                cur_p = p
            # Sheet cuối của bảng → lấy hết phần còn lại (chống lệch counter).
            max_rows = None if idx == total - 1 else clen
            _build_table_sheet(ws, job, p, row_iter, max_rows, idx, total,
                               summary_in_file=(pi == 0))
            if idx == 0 and pi == 0:
                titles[p.name] = title

        if summary_ws is not None:
            _build_summary(summary_ws, job, titles, first_part)

        buf = io.BytesIO()
        wb.save(buf)
        suffix = "" if n_parts == 1 else f"_part{pi + 1}"
        files.append((f"compare_{job_tag}{suffix}.xlsx", buf.getvalue()))

    zbuf = io.BytesIO()
    with zipfile.ZipFile(zbuf, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, data in files:
            zf.writestr(fname, data)
    return zbuf.getvalue()


# --------------------------------------------------------------------------- #
# CSV (phẳng — 1 dòng / phát hiện)
# --------------------------------------------------------------------------- #
def build_csv(job) -> str:
    buf = io.StringIO()
    buf.write("﻿")  # BOM để Excel mở UTF-8 đúng tiếng Việt
    w = csv.writer(buf)
    w.writerow([
        "table", "mode", "status", "finding",
        "id_or_key", "column", f"value_{job.label_a}", f"value_{job.label_b}",
        "detail",
    ])
    # Kết quả tìm từ khóa (dò ở DB B) — ghi lên đầu file cho dễ thấy.
    for h in getattr(job, "keyword_hits", []) or []:
        detail = h.query
        if h.error:
            detail = (detail + " | " if detail else "") + f"error={h.error}"
        w.writerow([
            h.table, "", "keyword_search",
            f"{_KIND_LABEL.get(h.kind, h.kind)}: {h.keyword}",
            ", ".join(str(i) for i in h.sample_ids),
            _fmt_cols_list(h.columns),
            "", h.match_count, detail,
        ])
    for p in job.progress.values():
        w.writerow([
            p.name, p.mode, p.status, "summary", "", "", p.count_a, p.count_b,
            f"delta={p.count_delta}; only_a={p.only_in_a}; only_b={p.only_in_b}; "
            f"value_mismatch={p.value_mismatch}"
            + (f"; note={p.note}" if p.note else "")
            + (f"; error={p.error}" if p.error else ""),
        ])
        if p.mode != "count-only":  # stream full từ spill
            pair = job.pair_key()
            for key in spill.iter_records(pair, p.name, "only_a"):
                w.writerow([p.name, p.mode, p.status, "only_in_a", key, "", "", "", ""])
            for key in spill.iter_records(pair, p.name, "only_b"):
                w.writerow([p.name, p.mode, p.status, "only_in_b", key, "", "", "", ""])
            for rec in spill.iter_records(pair, p.name, "mismatch_detail"):
                rid = rec.get("id")
                diffs = rec.get("diffs") or []
                if diffs:
                    for d in diffs:
                        w.writerow([
                            p.name, p.mode, p.status, "value_mismatch",
                            rid, d.get("col"), d.get("a"), d.get("b"), "",
                        ])
                else:
                    w.writerow([p.name, p.mode, p.status, "value_mismatch", rid, "", "", "", ""])

        # Schema + constraint diff (lấy từ schema_tables)
        sd = (getattr(job, "schema_tables", {}) or {}).get(p.name)
        if sd:
            for c in getattr(sd, "columns", []):
                w.writerow([
                    p.name, p.mode, p.status, f"column_{c.kind}",
                    "", c.name, c.detail_a or "", c.detail_b or "", "",
                ])
            for c in getattr(sd, "constraints", []):
                w.writerow([
                    p.name, p.mode, p.status, f"constraint_{c.kind}",
                    "", c.name, c.def_a or "", c.def_b or "", c.type_label,
                ])
    return buf.getvalue()
